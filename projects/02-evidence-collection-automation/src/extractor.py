"""
Minimal file text extractor for evidence uploads.
Reuses the extraction logic from Project 1's assessor.py.
"""

from pathlib import Path


def extract_file_text(file_path: str) -> str:
    """Extract text from any supported file type."""
    path = Path(file_path)
    suffix = path.suffix.lower()

    try:
        if suffix == ".pdf":
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                return "\n".join(
                    (p.extract_text() or "") for p in pdf.pages
                )
        elif suffix == ".csv":
            import pandas as pd
            df = pd.read_csv(path)
            return df.to_string()
        elif suffix == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"Sheet: {sheet_name}")
                for row in ws.iter_rows(values_only=True):
                    parts.append(" | ".join(str(c) if c else "" for c in row))
            return "\n".join(parts[:200])
        elif suffix in (".docx", ".doc"):
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)
        elif suffix in (".png", ".jpg", ".jpeg", ".gif", ".bmp"):
            from PIL import Image
            img = Image.open(path)
            return f"[Image: {path.name}, {img.size[0]}x{img.size[1]}, {img.mode}]"
        else:
            try:
                return path.read_text(encoding="utf-8")
            except UnicodeDecodeError:
                return f"[Binary file: {path.name} ({path.stat().st_size} bytes)]"
    except Exception as e:
        return f"[Extraction failed for {path.name}: {e}]"
